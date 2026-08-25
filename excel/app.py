import streamlit as st
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import io
import pytesseract
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
from pdf2image import convert_from_bytes

# --- 1. Helper Parsing Functions ---
def is_company_name(line):
    if len(line) < 3 or len(line) > 60: return False
    if '@' in line or 'www.' in line.lower(): return False
    kws = ['off', 'res', 'fax', 'mobile', 'email', 'contact', 'partner', 'company', 'deals', 'size', 'banker', 'proprietor', 'director', 'web', 'qbc']
    if any(line.lower().startswith(k) for k in kws): return False
    alphas = [c for c in line if c.isalpha()]
    if not alphas: return False
    return (len([c for c in alphas if c.isupper()]) / len(alphas)) > 0.75

def extract_numbers_regex(patterns, text):
    nums = []
    for pattern in patterns:
        for match in re.finditer(pattern, text, re.IGNORECASE):
            substring = text[match.end():]
            match_nums = re.match(r'^([\s\d,\/]+)', substring)
            if match_nums:
                n_str = match_nums.group(1)
                parts = re.split(r'[,/]', n_str)
                nums.extend([p.strip() for p in parts if any(c.isdigit() for c in p) and len(p.strip()) >= 4])
    return list(set(nums))

def get_safe(lst, idx, default=""):
    return lst[idx] if len(lst) > idx else default

# --- 2. Main Extraction and Parsing Logic ---
def parse_directory_text(raw_text):
    lines = [l.strip() for l in raw_text.split('\n') if l.strip()]
    companies = []
    current = None

    for line in lines:
        if line.isdigit() or "DIAMOND HANDBOOK" in line or "DIAMOND WORLD" in line or "The intermationul" in line or "57 FACETS" in line:
            continue
        if is_company_name(line):
            if current: companies.append(current)
            current = {"name": line, "lines": []}
        elif current is not None:
            current["lines"].append(line)
    if current: companies.append(current)

    parsed_data = []

    for c in companies:
        block = " ".join(c["lines"])
        
        # 10-digit mobile numbers
        mobiles_raw = re.findall(r'(?:\+91[\s\-]?)?([5-9]\d{9})\b', re.sub(r'[\s\-](?=\d)', '', block))
        mobiles = list(dict.fromkeys(mobiles_raw))
        
        emails = list(set(re.findall(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', block)))
        webs = list(set(re.findall(r'www\.[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', block.lower())))
        
        offices = extract_numbers_regex([r'\bOff[\.\:]?\b', r'\bO\.\b', r'Off\s'], block)
        qbcs = extract_numbers_regex([r'\bQbc[\.\:]?\b', r'\bQbo[\.\:]?\b', r'\bQu[\.\:]?\b', r'\bQ\.\b', r'\bCoc[\.\:]?\b'], block)
        ress = extract_numbers_regex([r'\bRes[\.\:]?\b', r'\bRe[\.\.]?\b'], block)

        address_lines = []
        for l in c["lines"]:
            l_lower = l.lower()
            if any(l_lower.startswith(k) for k in ['off', 'res', 'fax', 'mobile', 'email', 'contact', 'partner', 'company type', 'deals', 'size', 'banker', 'proprietor', 'director', 'qbc', 'web']):
                break
            address_lines.append(l)
        
        address = " ".join(address_lines)
        address = re.sub(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', '', address)
        address = re.sub(r'www\.[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', '', address).strip()
        
        city = ""
        for cty in ["Mumbai", "Surat", "Chandigarh", "New Delhi", "Delhi"]:
            if cty.lower() in address.lower(): city = cty; break
                
        partners, contacts, deals_in, size = [], [], [], []
        current_field = None
        
        for l in c["lines"]:
            l_lower = l.lower()
            
            if 'contact' in l_lower and ('person' in l_lower or 'porcon' in l_lower):
                val = re.sub(r'(?i).*contact\s*p[eo]r[cs]on\s*:?\s*', '', l).strip()
                if val: contacts.extend([x.strip() for x in val.split(',')])
                current_field = 'contact'
                continue
                
            if l_lower.startswith('partner') or l_lower.startswith('proprietor') or l_lower.startswith('director'):
                val = l.split(':', 1)[-1].strip()
                if val and val.lower() not in ['partner', 'partners', 'proprietor', 'director']:
                    partners.extend([x.strip() for x in val.split(',')])
                current_field = 'partner'
                continue
                
            if l_lower.startswith('deals in'):
                val = l.split(':', 1)[-1].strip()
                if val and val.lower() != 'deals in': deals_in.append(val)
                current_field = 'deals'
                continue
                
            if l_lower.startswith('size'):
                val = l.split(':', 1)[-1].strip()
                if val and val.lower() != 'size': size.append(val)
                current_field = 'size'
                continue
                
            if any(l_lower.startswith(k) for k in ['banker', 'company', 'off', 'res', 'fax', 'mobile', 'email', 'web', 'qbc']):
                current_field = None
                continue
                
            if current_field == 'contact':
                if not any(char.isdigit() for char in l) and len(l) > 2 and not '@' in l:
                    contacts.extend([x.strip() for x in l.split(',')])
            elif current_field == 'partner':
                if not any(char.isdigit() for char in l) and len(l) > 2 and not '@' in l:
                    partners.extend([x.strip() for x in l.split(',')])
            elif current_field == 'deals':
                if len(l) > 2 and not '@' in l and not any(l_lower.startswith(k) for k in ['banker', 'company', 'a.', 'b.', 'c.', 'd.']):
                    deals_in.append(l.strip())
            elif current_field == 'size':
                if len(l) > 2 and not '@' in l and not any(l_lower.startswith(k) for k in ['banker', 'company', 'a.', 'b.', 'c.', 'd.']):
                    size.append(l.strip())
                    
        partners = [p for p in partners if p and len(p)>2 and not any(c.isdigit() for c in p)]
        contacts = [re.sub(r'(?i)^contact\s*p[eo]r[cs]on\s*:?\s*', '', c).strip() for c in contacts]
        contacts = [c for c in contacts if c and len(c)>2 and not any(ch.isdigit() for ch in c)]
        
        parsed_data.append({
            "Company Name": c["name"],
            "Address": address,
            "City": city,
            "Proprietor/Partner 1": get_safe(partners, 0),
            "Proprietor/Partner 2": get_safe(partners, 1),
            "Proprietor/Partner 3": get_safe(partners, 2),
            "Contact Person 1": get_safe(contacts, 0),
            "Contact Person 2": get_safe(contacts, 1),
            "Contact Person 3": get_safe(contacts, 2),
            "Off 1": get_safe(offices, 0),
            "Off 2": get_safe(offices, 1),
            "Off 3": get_safe(offices, 2),
            "Qbc 1": get_safe(qbcs, 0),
            "Qbc 2": get_safe(qbcs, 1),
            "Qbc 3": get_safe(qbcs, 2),
            "Mobile 1": get_safe(mobiles, 0),
            "Mobile 2": get_safe(mobiles, 1),
            "Mobile 3": get_safe(mobiles, 2),
            "Res 1": get_safe(ress, 0),
            "Res 2": get_safe(ress, 1),
            "Res 3": get_safe(ress, 2),
            "Email 1": get_safe(emails, 0),
            "Email 2": get_safe(emails, 1),
            "Email 3": get_safe(emails, 2),
            "Size": ", ".join(size),
            "Deals in": ", ".join(deals_in),
            "Website": get_safe(webs, 0)
        })

    return pd.DataFrame(parsed_data)

# --- 3. Excel Generation Logic ---
def style_worksheet(ws):
    """Helper function to apply styling to any worksheet."""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="333333")
    alt_fill = PatternFill("solid", fgColor="F9F9F9")
    border = Border(left=Side(style='thin', color='D3D3D3'), right=Side(style='thin', color='D3D3D3'), top=Side(style='thin', color='D3D3D3'), bottom=Side(style='thin', color='D3D3D3'))

    # Style Header
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # Style Rows & Apply Alternating Colors
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column), start=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # Set Column Widths
    for col_letter in ['A', 'B', 'V', 'W', 'X', 'Y', 'Z', 'AA']: ws.column_dimensions[col_letter].width = 30
    for col_letter in ['C', 'D', 'E', 'F', 'G', 'H', 'I']: ws.column_dimensions[col_letter].width = 20
    for col_letter in ['J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U']: ws.column_dimensions[col_letter].width = 15

    ws.freeze_panes = "A2"

def create_styled_excel(df):
    wb = Workbook()
    
    # -- Sheet 1: Main Data --
    ws_main = wb.active
    ws_main.title = "Enhanced Directory Data"
    
    ws_main.append(list(df.columns))
    for row in dataframe_to_rows(df, index=False, header=False): 
        ws_main.append(row)
    style_worksheet(ws_main)

    # -- Sheet 2: Duplicate Companies Only --
    # Temporarily normalize names to catch case-insensitive duplicates
    df['normalized_name'] = df['Company Name'].str.strip().str.upper()
    # Filter for duplicates (keep=False ensures we keep ALL instances of the duplicate)
    duplicates_df = df[df.duplicated(subset=['normalized_name'], keep=False)].drop(columns=['normalized_name'])
    df = df.drop(columns=['normalized_name']) # Cleanup main df just in case
    
    # Sort them so matching duplicate branches group together visually
    duplicates_df = duplicates_df.sort_values(by=['Company Name'])
    
    # Only create the duplicates sheet if duplicates exist
    if not duplicates_df.empty:
        ws_dupes = wb.create_sheet("Duplicate Companies")
        ws_dupes.append(list(duplicates_df.columns))
        for row in dataframe_to_rows(duplicates_df, index=False, header=False): 
            ws_dupes.append(row)
        style_worksheet(ws_dupes)

    # Save to buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# --- 4. Streamlit UI ---
st.set_page_config(page_title="PDF to Directory Excel", layout="wide")
st.title("📄 Scanned PDF Directory to Excel Converter")
st.markdown("Upload your scanned directory PDF. The app will extract the text using OCR and generate a neatly formatted Excel file.")

uploaded_file = st.file_uploader("Upload Scanned PDF", type=["pdf"])

if uploaded_file is not None:
    if st.button("Extract and Convert"):
        with st.spinner("Converting PDF pages to images & running OCR (this may take a minute)..."):
            try:
                # IMPORTANT: If you are running locally on Windows and need the Tesseract Path, keep this line. 
                # If deploying to Streamlit Cloud, comment this out!
                pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

                # Convert PDF to images
                images = convert_from_bytes(uploaded_file.read())
                
                # Run OCR on each image
                extracted_text = ""
                for i, image in enumerate(images):
                    text = pytesseract.image_to_string(image)
                    extracted_text += text + "\n"
                
                if not extracted_text.strip():
                    st.error("No text could be extracted from the PDF. Please ensure the scan quality is clear.")
                else:
                    st.success(f"Successfully extracted text from {len(images)} pages. Parsing data now...")
                    
                    with st.spinner("Structuring and formatting data into Excel..."):
                        # Parse the text
                        df = parse_directory_text(extracted_text)
                        
                        # Generate Excel
                        excel_buffer = create_styled_excel(df)
                        
                        st.success(f"Processing Complete! Successfully parsed {len(df)} companies.")
                        
                        st.download_button(
                            label="📥 Download Formatted Excel File",
                            data=excel_buffer,
                            file_name="Structured_Directory_Data.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                        
                        st.subheader("Data Preview (Main Table)")
                        st.dataframe(df.head())
                        
            except Exception as e:
                st.error(f"An error occurred during processing: {e}")